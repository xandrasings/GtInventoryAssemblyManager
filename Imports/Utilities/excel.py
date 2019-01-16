from .const import *
from .fileManagement import *
from .sys import *
from ..Classes.Component import *
from ..Classes.Product import *
from ..Classes.Task import *

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.page import PageMargins

import os


def loadWorkbook(fileType, filePath, readOnly = False):
	workBook = load_workbook(filePath, readOnly)
	assertExpectedWorkSheets(fileType, workBook)
	return workBook


def openSheet(workBook, sheetType):
	sheet = workBook[sheetType]
	verifySheetFormatting(sheet, sheetType)
	return sheet


def verifySheetFormatting(sheet, sheetType):
	# TODO assert worksheet dimensions from ws.calculate_dimension() sheet.get_highest_column()
	# TODO assert column headers are as expected
	# TODO use regex to assert expected formatting of rows
	pass


def assertExpectedWorkSheets(fileType, workBook):
	for sheet in FILE_SHEETS[fileType]:
		if not sheet in workBook.sheetnames:
			print('Selected task file is missing ' + sheet + ' worksheet.')
			quit()


def generateTaskList():
	taskList = []

	workBook = loadWorkbook(TASKS, selectTaskFile(), True)
	taskSheet = workBook[TASK_FILE_TASK_SHEET]
	productDictionary = generateProductDictionary(workBook)

	for i in range(2, taskSheet.max_row + 1):
		product = str(taskSheet.cell(row = i, column = TASK_LIST_PRODUCT_COLUMN).value)
		if product == "None":
			break
		quantity = taskSheet.cell(row = i, column = TASK_LIST_QUANTITY_COLUMN).value

		taskList.append(Task(i - 1, productDictionary[product], quantity))

	return taskList


def generateProductDictionary(workBook):
	productDictionary = generateEmptyProductDictionary(workBook)
	return populateProductDictionary(workBook, productDictionary)


def generateEmptyProductDictionary(workBook):
	productSheet = openSheet(workBook, TASK_FILE_PRODUCTS_SHEET)
	productDictionary = {}

	for i in range(2, productSheet.max_row + 1):
		name = str(productSheet.cell(row = i, column = PRODUCT_LIST_PRODUCT_COLUMN).value)
		timeEstimate = productSheet.cell(row = i, column = PRODUCT_LIST_TIME_ESTIMATE_COLUMN).value

		if name not in productDictionary.keys():
			productDictionary[name] = Product(name, timeEstimate)

	return productDictionary


def populateProductDictionary(workBook, productDictionary):
	componentSheet = openSheet(workBook, TASK_FILE_COMPONENTS_SHEET)

	for i in range(2, componentSheet.max_row + 1):
		product = str(componentSheet.cell(row = i, column = COMPONENT_LIST_PRODUCT_COLUMN).value)
		part = str(componentSheet.cell(row = i, column = COMPONENT_LIST_PART_COLUMN).value)
		quantity = componentSheet.cell(row = i, column = COMPONENT_LIST_QUANTITY_COLUMN).value
		
		productDictionary[product].addComponent(Component(part, quantity))

	assertComponentMax(productDictionary)
	return productDictionary


def assertComponentMax(productDictionary):
	for product in productDictionary:
		if len(productDictionary[product].getComponents()) > PRODUCT_COMPONENT_LIST_MAX:
			print('Product ' + product + ' has over the max (' + str(PRODUCT_COMPONENT_LIST_MAX) + ') of assembly components.')
			quit()


def generateAssemblyOrderFile(taskList):
	createAssemblyOrderFile()
	try:
		workBook = loadWorkbook(TEMPLATE, ASSEMBLY_ORDER_FILE_PATH)
		fileName = generateAssemblyOrderFileName()
		populateAssemblyOrderFile(workBook, taskList, fileName)
		saveAndRenameWorkbook(workBook, fileName)
		print('Successfully generated assembly order file: ' + fileName)
	except Exception as e:
		print('Something went wrong! Cancelling process.')
		print('\n\n\nerror detail:')
		print(e)
		print('\n\n\n')
		deleteAssemblyOrderFile()
		quit()


def saveAndRenameWorkbook(workBook, fileName):
		workBook.save(ASSEMBLY_ORDER_FILE_PATH)
		return renameAssemblyOrderFile(fileName)


def populateAssemblyOrderFile(workBook, taskList, fileName):
	populateCopyrightPage(workBook)
	populateAssemblyOrderPages(workBook, taskList, fileName)


def populateCopyrightPage(workBook):
	copyrightSheet = workBook[TEMPLATE_FILE_COPYRIGHT_SHEET]
	addImage(copyrightSheet, VERTEX_IMAGE_FILE_PATH, 'B3')


def populateAssemblyOrderPages(workBook, taskList, fileName):
	templateSheet = workBook[TEMPLATE_FILE_TEMPLATE_SHEET]
	taskCount = len(taskList)

	for i in range (0, taskCount):
		product = taskList[i].getProduct()
		newSheet = workBook.copy_worksheet(templateSheet)
		newSheet.title = generateSheetName(i, product.getName())
		pageMargins = PageMargins(top = margin[TOP], bottom = margin[BOTTOM], left = margin[LEFT], right = margin[RIGHT], header = margin[HEADER], footer = margin[FOOTER])

		populateCell(newSheet, FRACTION, generateFraction(i, taskCount))
		populateCell(newSheet, PRODUCT, product.getName())
		populateCell(newSheet, QUANTITY, taskList[i].getQuantity())
		populateCell(newSheet, TIME_ESTIMATE, taskList[i].getTimeEstimate())
		populateCell(newSheet, FILE_NAME, fileName)

		currentRow = PRODUCT_COMPONENT_LIST_START_ROW
		for component in taskList[i].getComponents():
			populateCell(newSheet, PART, component.getPart(), currentRow)
			populateCell(newSheet, PART_QUANTITY, component.getQuantity(), currentRow)
			currentRow = currentRow + 1
		formatAssemblyOrderSheet(newSheet, pageMargins)
	workBook.remove(templateSheet)
	

def generateSheetName(counter, product):
	return str(counter + 1) + " " + product[:27]


def generateFraction(counter, total):
	return str(counter + 1) + "/" + str(total)


def populateCell(sheet, dataElementType, value, currentRow = None):
	sheet.cell(row = calculateDataElementRow(dataElementType, currentRow), column = dataElementInfo[dataElementType][COLUMN]).value = value;


def calculateDataElementRow(dataElementType, currentRow):
	if dataElementType not in [PART, PART_QUANTITY]:
		currentRow = dataElementInfo[dataElementType][ROW]

	if currentRow == None:
		print('Encountered error calculating target row for ' + dataElementType)
		quit()

	return currentRow


def addImage(sheet, imagePath, targetCell):
	image = Image(imagePath)
	sheet.add_image(image, targetCell)


def addProductAndQuantityUnderline(sheet):
	for dataElementType in [PRODUCT, QUANTITY]:
		addUnderline(sheet, dataElementInfo[dataElementType][ROW] + dataElementInfo[dataElementType][ROW_BORDER_OFFSET], dataElementInfo[dataElementType][COLUMN] + dataElementInfo[dataElementType][COLUMN_BORDER_OFFSET])


def addUnderline(sheet, targetRow, targetColumn):
	border = Border(bottom = Side(style='thin', color='00000000'))
	sheet.cell(row = targetRow, column = targetColumn).border = border


def addSingleRowBoxBorder(sheet, targetRow, start, stop):
	border = Border(top = Side(style='thin', color='00000000'), bottom = Side(style='thin', color='00000000'))
	leftEndBorder = Border(top = Side(style='thin', color='00000000'), bottom = Side(style='thin', color='00000000'), left = Side(style='thin', color='00000000'))
	rightEndBorder = Border(top = Side(style='thin', color='00000000'), bottom = Side(style='thin', color='00000000'), right = Side(style='thin', color='00000000'))
	
	for targetColumn in range(start + 1, stop - 1):
		sheet.cell(row = targetRow, column = targetColumn).border = border

	sheet.cell(row = targetRow, column = start).border = leftEndBorder
	sheet.cell(row = targetRow, column = stop - 1).border = rightEndBorder


def leftAlignAssemblyOrderNumbers(sheet):
	for targetRow in range (PRODUCT_COMPONENT_LIST_START_ROW, PRODUCT_COMPONENT_LIST_START_ROW + PRODUCT_COMPONENT_LIST_MAX):
		sheet.cell(row = targetRow, column = dataElementInfo[PART_QUANTITY][COLUMN]).alignment = Alignment(horizontal = 'left')
		sheet.cell(row = targetRow, column = dataElementInfo[ORDER_QUANTITY][COLUMN]).alignment = Alignment(horizontal = 'left')


def assignPageMargins(sheet, pageMargins):
	sheet.page_margins = pageMargins
	sheet.print_options.horizontalCentered = True


def assignPrintArea(sheet):
	sheet.print_area = PRINT_AREA


def formatAssemblyOrderSheet(sheet, pageMargins):
	addImage(sheet, GLACIER_TEK_IMAGE_FILE_PATH, 'A1')
	addProductAndQuantityUnderline(sheet)
	addSingleRowBoxBorder(sheet, PRODUCT_COMPONENT_LIST_START_ROW - 1, PRODUCT_COMPONENT_LIST_START_COLUMN, PRODUCT_COMPONENT_LIST_END_COLUMN + 1)
	leftAlignAssemblyOrderNumbers(sheet)
	assignPageMargins(sheet, pageMargins)
	assignPrintArea(sheet)
